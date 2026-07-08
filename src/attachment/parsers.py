"""
附件解析器
支持 PDF、Word、Excel、纯文本、压缩包
"""
import logging
import zipfile
import os
from pathlib import Path

logger = logging.getLogger(__name__)


def parse_pdf(file_path: str) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except ImportError:
        logger.warning("pypdf 未安装")
        return f"[PDF: {Path(file_path).name}]"
    except Exception as e:
        logger.error(f"PDF 解析失败: {e}")
        return ""


def parse_docx(file_path: str) -> str:
    try:
        from docx import Document
        doc = Document(file_path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        logger.warning("python-docx 未安装")
        return f"[Word: {Path(file_path).name}]"
    except Exception as e:
        logger.error(f"Word 解析失败: {e}")
        return ""


def parse_excel(file_path: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        all_texts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_texts.append(f"[工作表: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(c) if c is not None else "" for c in row)
                if row_text.strip():
                    all_texts.append(row_text)
        wb.close()
        return "\n".join(all_texts)
    except ImportError:
        logger.warning("openpyxl 未安装")
        return f"[Excel: {Path(file_path).name}]"
    except Exception as e:
        logger.error(f"Excel 解析失败: {e}")
        return ""


def parse_text(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as e:
        logger.error(f"文本读取失败: {e}")
        return ""


def extract_and_parse_zip(file_path: str, extract_dir: str) -> str:
    try:
        dest = Path(extract_dir) / f"_unzip_{Path(file_path).stem}"
        dest.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(file_path, "r") as zf:
            zf.extractall(dest)
        texts = []
        for root, _, files in os.walk(dest):
            for f in files:
                fpath = os.path.join(root, f)
                ext = Path(f).suffix.lower()
                if ext == ".pdf": texts.append(parse_pdf(fpath))
                elif ext in (".docx", ".doc"): texts.append(parse_docx(fpath))
                elif ext in (".xlsx", ".xls", ".csv"): texts.append(parse_excel(fpath))
                elif ext == ".txt": texts.append(parse_text(fpath))
        return "\n---\n".join(texts)
    except Exception as e:
        logger.error(f"压缩包解析失败: {e}")
        return ""
